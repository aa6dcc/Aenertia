import serial
import json
import os
from openpyxl import Workbook, load_workbook

# === CONFIGURATION ===
SERIAL_PORT = 'COM4'      # Change to your actual serial port
BAUD_RATE = 115200          # Match the baud rate with your device
EXCEL_FILE = 'PowerLog.xlsx'
SHEET_NAME = 'Sheet1'
# ======================
MyString = 'PM: {"voltage":3,"current_motor":6,"current_board":9} {"voltage":4,"current_motor":7,"current_board":10} {"voltage":5,"current_motor":8,"current_board":1} {"voltage":6,"current_motor":9,"current_board":2} {"voltage":7,"current_motor":10,"current_board":3} {"voltage":8,"current_motor":1,"current_board":4} {"voltage":9,"current_motor":2,"current_board":5} {"voltage":10,"current_motor":3,"current_board":6} {"voltage":1,"current_motor":4,"current_board":7} {"voltage":2,"current_motor":5,"current_board":8} {"voltage":3,"current_motor":6,"current_board":9} {"voltage":4,"current_motor":7,"current_board":10} {"voltage":5,"current_motor":8,"current_board":1} {"voltage":6,"current_motor":9,"current_board":2} {"voltage":7,"current_motor":10,"current_board":3} {"voltage":8,"current_motor":1,"current_board":4} {"voltage":9,"current_motor":2,"current_board":5} {"voltage":10,"current_motor":3,"current_board":6} {"voltage":1,"current_motor":4,"current_board":7} {"voltage":2,"current_motor":5,"current_board":8}'

def initialize_excel_file():
    if not os.path.exists(EXCEL_FILE):
        print("Initialized File")
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["Voltage", "Current_Motor", "Current_Board"])
        wb.save(EXCEL_FILE)

def append_to_excel(voltage, current_motor, current_board):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    ws.append([voltage, current_motor, current_board])
    wb.save(EXCEL_FILE)


if MyString[0:3] == "PM:":
    MyList = MyString.split();
    for i in range(1, len(MyList)):
        data = json.loads(MyList[i])
        voltage = data.get("voltage")
        current_motor = data.get("current_motor")
        current_board = data.get("current_board")
        if voltage is not None and current_motor is not None and current_board is not None:
            append_to_excel(voltage, current_motor, current_board)
            print(f"Logged: {voltage}, {current_motor}, {current_board}")
        else:
            print("Missing one or more required keys in JSON.")



# def main():
#     initialize_excel_file()
#     with serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=1) as ser:
#         print("Listening for serial input...")
#         while True:
#             try:
#                 line = ser.readline().decode('utf-8').strip()
#                 if not line:
#                     continue
#                 else:
#                     print(line)

#                 if line[0:3] == "PM:":
#                     line = line.split()[1];
#                     data = json.loads(line)
#                     voltage = data.get("voltage")
#                     current_motor = data.get("current_motor")
#                     current_board = data.get("current_board")
#                 if voltage is not None and current_motor is not None and current_board is not None:
#                     append_to_excel(voltage, current_motor, current_board)
#                     print(f"Logged: {voltage}, {current_motor}, {current_board}")
#                 else:
#                     print("Missing one or more required keys in JSON.")

#             except json.JSONDecodeError:
#                 print("Invalid JSON received.")
#             except Exception as e:
#                 print(f"Error: {e}")

# if __name__ == "__main__":
#     main()